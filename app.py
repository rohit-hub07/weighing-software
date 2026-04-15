import os
import random
import re
import threading
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, ttk
from urllib.parse import unquote

import requests

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise ImportError("openpyxl is required. Install it using: pip install openpyxl")

try:
    from twilio.rest import Client
except ImportError:
    raise ImportError("twilio is required. Install it using: pip install twilio")


EXCEL_FILE = "weighment_data.xlsx"
HEADERS = [
    "Serial No",
    "Vehicle No",
    "Date",
    "Time",
    "Challan",
    "Customer Code",
    "Customer Name",
    "Product Code",
    "Product Name",
    "Source Code",
    "Source Name",
    "Destination Code",
    "Destination Name",
    "Transporter Code",
    "Transporter Name",
    "Gross Weight",
    "Tare Weight",
    "Net Weight",
]


class WeighmentApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Weighment Section")
        self.root.geometry("1080x760")
        self.root.minsize(900, 620)

        self.current_weight = 0
        self.gross_weight = None
        self.tare_weight = None
        self.net_weight = None

        self.serial_no_var = tk.StringVar(value=str(self._get_next_serial_no()))
        self.vehicle_no_var = tk.StringVar()
        self.date_var = tk.StringVar()
        self.time_var = tk.StringVar()
        self.challan_var = tk.StringVar()
        self.customer_code_var = tk.StringVar()
        self.customer_name_var = tk.StringVar()
        self.product_code_var = tk.StringVar()
        self.product_name_var = tk.StringVar()
        self.source_code_var = tk.StringVar()
        self.source_name_var = tk.StringVar()
        self.desti_code_var = tk.StringVar()
        self.desti_name_var = tk.StringVar()
        self.transporter_code_var = tk.StringVar()
        self.transporter_name_var = tk.StringVar()

        self.twilio_sid_var = tk.StringVar(value=os.getenv("TWILIO_ACCOUNT_SID", ""))
        self.twilio_token_var = tk.StringVar(value=os.getenv("TWILIO_AUTH_TOKEN", ""))
        self.twilio_sms_from_var = tk.StringVar(value=os.getenv("TWILIO_SMS_FROM", ""))
        self.twilio_whatsapp_from_var = tk.StringVar(value=os.getenv("TWILIO_WHATSAPP_FROM", ""))
        self.sms_to_var = tk.StringVar(value=os.getenv("SMS_TO", ""))
        self.whatsapp_to_var = tk.StringVar(value=os.getenv("WHATSAPP_TO", ""))
        self.telegram_bot_token_var = tk.StringVar(value=os.getenv("TELEGRAM_BOT_TOKEN", ""))
        self.telegram_chat_id_var = tk.StringVar(value=os.getenv("TELEGRAM_CHAT_ID", ""))
        self.message_status_var = tk.StringVar(value="Messaging: Ready")

        self.live_weight_var = tk.StringVar(value="00000 kg")
        self.gross_var = tk.StringVar(value="-")
        self.tare_var = tk.StringVar(value="-")
        self.net_var = tk.StringVar(value="-")

        self._ensure_workbook()
        self._build_ui()
        self._update_datetime()
        self.generate_weight()

    def _build_ui(self) -> None:
        outer = ttk.Frame(self.root)
        outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer, highlightthickness=0)
        vscroll = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)

        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        container = ttk.Frame(canvas, padding=16)
        window_id = canvas.create_window((0, 0), window=container, anchor="nw")

        def _sync_scroll_region(_event=None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfigure(window_id, width=canvas.winfo_width())

        container.bind("<Configure>", _sync_scroll_region)
        canvas.bind("<Configure>", _sync_scroll_region)

        def _on_mousewheel(event) -> None:
            canvas.yview_scroll(int(-event.delta / 120), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        input_card = ttk.LabelFrame(container, text="Weighment Entry", padding=12)
        input_card.pack(fill="x")

        ttk.Label(input_card, text="Serial No").grid(row=0, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.serial_no_var, state="readonly", width=20).grid(
            row=0, column=1, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="Vehicle No").grid(row=1, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.vehicle_no_var, width=24).grid(
            row=1, column=1, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="Date").grid(row=0, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.date_var, state="readonly", width=20).grid(
            row=0, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="Time").grid(row=1, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.time_var, state="readonly", width=20).grid(
            row=1, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="CHALLAN").grid(row=2, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.challan_var, width=24).grid(
            row=2, column=1, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="CUSTOMER Code").grid(row=3, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.customer_code_var, width=24).grid(
            row=3, column=1, sticky="w", padx=6, pady=6
        )
        ttk.Label(input_card, text="CUSTOMER Name").grid(row=3, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.customer_name_var, width=30).grid(
            row=3, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="PRODUCT Code").grid(row=4, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.product_code_var, width=24).grid(
            row=4, column=1, sticky="w", padx=6, pady=6
        )
        ttk.Label(input_card, text="PRODUCT Name").grid(row=4, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.product_name_var, width=30).grid(
            row=4, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="SOURCE Code").grid(row=5, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.source_code_var, width=24).grid(
            row=5, column=1, sticky="w", padx=6, pady=6
        )
        ttk.Label(input_card, text="SOURCE Name").grid(row=5, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.source_name_var, width=30).grid(
            row=5, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="DESTI Code").grid(row=6, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.desti_code_var, width=24).grid(
            row=6, column=1, sticky="w", padx=6, pady=6
        )
        ttk.Label(input_card, text="DESTI Name").grid(row=6, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.desti_name_var, width=30).grid(
            row=6, column=3, sticky="w", padx=6, pady=6
        )

        ttk.Label(input_card, text="TRANSPOTER Code").grid(row=7, column=0, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.transporter_code_var, width=24).grid(
            row=7, column=1, sticky="w", padx=6, pady=6
        )
        ttk.Label(input_card, text="TRANSPOTER Name").grid(row=7, column=2, sticky="w", padx=6, pady=6)
        ttk.Entry(input_card, textvariable=self.transporter_name_var, width=30).grid(
            row=7, column=3, sticky="w", padx=6, pady=6
        )

        display_frame = ttk.Frame(container, padding=(0, 14, 0, 10))
        display_frame.pack(fill="x")

        ttk.Label(display_frame, text="Live Weight", font=("Segoe UI", 12, "bold")).pack(anchor="w")
        tk.Label(
            display_frame,
            textvariable=self.live_weight_var,
            font=("Consolas", 36, "bold"),
            bg="black",
            fg="#28ff6d",
            padx=20,
            pady=14,
            relief="sunken",
            bd=4,
        ).pack(fill="x", pady=(8, 0))

        button_row = ttk.Frame(container, padding=(0, 12, 0, 8))
        button_row.pack(fill="x")

        ttk.Button(button_row, text="Gross Weight", command=self.capture_gross_weight).pack(
            side="left", padx=(0, 10)
        )
        ttk.Button(button_row, text="Tare Weight", command=self.capture_tare_weight).pack(
            side="left", padx=(0, 10)
        )
        ttk.Button(button_row, text="Net Weight", command=self.calculate_net_weight).pack(
            side="left", padx=(0, 10)
        )
        ttk.Button(button_row, text="Save", command=self.save_to_excel).pack(side="left", padx=(0, 10))
        ttk.Button(button_row, text="Clear", command=self.clear_fields).pack(side="left")

        result_card = ttk.LabelFrame(container, text="Captured Weights", padding=12)
        result_card.pack(fill="x", pady=(8, 0))

        ttk.Label(result_card, text="Gross Weight").grid(row=0, column=0, sticky="w", padx=6, pady=8)
        ttk.Label(result_card, textvariable=self.gross_var, font=("Segoe UI", 11, "bold")).grid(
            row=0, column=1, sticky="w", padx=6, pady=8
        )

        ttk.Label(result_card, text="Tare Weight").grid(row=1, column=0, sticky="w", padx=6, pady=8)
        ttk.Label(result_card, textvariable=self.tare_var, font=("Segoe UI", 11, "bold")).grid(
            row=1, column=1, sticky="w", padx=6, pady=8
        )

        ttk.Label(result_card, text="Net Weight").grid(row=2, column=0, sticky="w", padx=6, pady=8)
        ttk.Label(result_card, textvariable=self.net_var, font=("Segoe UI", 11, "bold")).grid(
            row=2, column=1, sticky="w", padx=6, pady=8
        )

        messaging_card = ttk.LabelFrame(container, text="Messaging", padding=12)
        messaging_card.pack(fill="x", pady=(10, 0))

        ttk.Label(messaging_card, text="Twilio SID").grid(row=0, column=0, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.twilio_sid_var, width=36).grid(
            row=0, column=1, sticky="w", padx=6, pady=5
        )
        ttk.Label(messaging_card, text="Twilio Token").grid(row=0, column=2, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.twilio_token_var, width=30, show="*").grid(
            row=0, column=3, sticky="w", padx=6, pady=5
        )

        ttk.Label(messaging_card, text="SMS From").grid(row=1, column=0, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.twilio_sms_from_var, width=36).grid(
            row=1, column=1, sticky="w", padx=6, pady=5
        )
        ttk.Label(messaging_card, text="SMS To").grid(row=1, column=2, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.sms_to_var, width=30).grid(
            row=1, column=3, sticky="w", padx=6, pady=5
        )

        ttk.Label(messaging_card, text="WhatsApp From").grid(row=2, column=0, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.twilio_whatsapp_from_var, width=36).grid(
            row=2, column=1, sticky="w", padx=6, pady=5
        )
        ttk.Label(messaging_card, text="WhatsApp To").grid(row=2, column=2, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.whatsapp_to_var, width=30).grid(
            row=2, column=3, sticky="w", padx=6, pady=5
        )

        ttk.Label(messaging_card, text="Telegram Bot Token").grid(row=3, column=0, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.telegram_bot_token_var, width=36, show="*").grid(
            row=3, column=1, sticky="w", padx=6, pady=5
        )
        ttk.Label(messaging_card, text="Telegram Chat ID").grid(row=3, column=2, sticky="w", padx=6, pady=5)
        ttk.Entry(messaging_card, textvariable=self.telegram_chat_id_var, width=30).grid(
            row=3, column=3, sticky="w", padx=6, pady=5
        )

        messaging_buttons = ttk.Frame(messaging_card)
        messaging_buttons.grid(row=4, column=0, columnspan=4, sticky="w", padx=6, pady=(10, 2))

        ttk.Button(messaging_buttons, text="Send SMS", command=self.send_sms).pack(side="left", padx=(0, 10))
        ttk.Button(messaging_buttons, text="Send WhatsApp", command=self.send_whatsapp).pack(
            side="left", padx=(0, 10)
        )
        ttk.Button(messaging_buttons, text="Send Telegram", command=self.send_telegram).pack(side="left")

        ttk.Label(messaging_card, textvariable=self.message_status_var).grid(
            row=5, column=0, columnspan=4, sticky="w", padx=6, pady=(6, 0)
        )

    def _update_datetime(self) -> None:
        now = datetime.now()
        self.date_var.set(now.strftime("%d-%m-%Y"))
        self.time_var.set(now.strftime("%H:%M:%S"))
        self.root.after(1000, self._update_datetime)

    def generate_weight(self) -> None:
        self.current_weight = random.randint(1000, 50000)
        self.live_weight_var.set(f"{self.current_weight:05d} kg")
        self.root.after(1000, self.generate_weight)

    def capture_gross_weight(self) -> None:
        self.gross_weight = self.current_weight
        self.gross_var.set(f"{self.gross_weight} kg")
        self.calculate_net_weight()

    def capture_tare_weight(self) -> None:
        self.tare_weight = self.current_weight
        self.tare_var.set(f"{self.tare_weight} kg")
        self.calculate_net_weight()

    def calculate_net_weight(self) -> None:
        if self.gross_weight is None or self.tare_weight is None:
            self.net_weight = None
            self.net_var.set("-")
            return

        self.net_weight = self.gross_weight - self.tare_weight
        self.net_var.set(f"{self.net_weight} kg")

    def _get_next_serial_no(self) -> int:
        if not os.path.exists(EXCEL_FILE):
            return 1

        try:
            workbook = load_workbook(EXCEL_FILE)
        except PermissionError:
            messagebox.showwarning(
                "Excel File In Use",
                "Close weighment_data.xlsx and restart the app to update missing column headers.",
            )
            return

        sheet = workbook.active

        if sheet.max_row <= 1:
            workbook.close()
            return 1

        last_serial = sheet.cell(row=sheet.max_row, column=1).value
        workbook.close()

        try:
            return int(last_serial) + 1
        except (TypeError, ValueError):
            return sheet.max_row

    def _ensure_workbook(self) -> None:
        if not os.path.exists(EXCEL_FILE):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(HEADERS)
            workbook.save(EXCEL_FILE)
            workbook.close()
            return

        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        changed = False
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            if cell.value != header:
                cell.value = header
                changed = True

        if changed:
            try:
                workbook.save(EXCEL_FILE)
            except PermissionError:
                messagebox.showwarning(
                    "Excel File In Use",
                    "Close weighment_data.xlsx and restart the app to update missing column headers.",
                )
        workbook.close()

    def _compose_message(self) -> str:
        self.calculate_net_weight()
        gross_text = str(self.gross_weight) if self.gross_weight is not None else "N/A"
        tare_text = str(self.tare_weight) if self.tare_weight is not None else "N/A"
        net_text = str(self.net_weight) if self.net_weight is not None else "N/A"

        return (
            f"Weighment Slip\n"
            f"Serial: {self.serial_no_var.get()}\n"
            f"Vehicle: {self.vehicle_no_var.get().strip() or 'N/A'}\n"
            f"Date: {self.date_var.get()}  Time: {self.time_var.get()}\n"
            f"Challan: {self.challan_var.get().strip() or 'N/A'}\n"
            f"Customer: {self.customer_name_var.get().strip() or 'N/A'} ({self.customer_code_var.get().strip() or 'N/A'})\n"
            f"Product: {self.product_name_var.get().strip() or 'N/A'} ({self.product_code_var.get().strip() or 'N/A'})\n"
            f"Source: {self.source_name_var.get().strip() or 'N/A'} ({self.source_code_var.get().strip() or 'N/A'})\n"
            f"Destination: {self.desti_name_var.get().strip() or 'N/A'} ({self.desti_code_var.get().strip() or 'N/A'})\n"
            f"Transporter: {self.transporter_name_var.get().strip() or 'N/A'} ({self.transporter_code_var.get().strip() or 'N/A'})\n"
            f"Gross: {gross_text} kg\n"
            f"Tare: {tare_text} kg\n"
            f"Net: {net_text} kg"
        )

    def _set_status(self, text: str) -> None:
        self.root.after(0, lambda: self.message_status_var.set(text))

    def _show_info(self, title: str, text: str) -> None:
        self.root.after(0, lambda: messagebox.showinfo(title, text))

    def _show_error(self, title: str, text: str) -> None:
        self.root.after(0, lambda: messagebox.showerror(title, text))

    def _run_in_background(self, send_fn, channel_name: str) -> None:
        message_text = self._compose_message()

        def _task() -> None:
            try:
                self._set_status(f"Messaging: Sending {channel_name}...")
                send_fn(message_text)
                self._set_status(f"Messaging: {channel_name} sent successfully")
                self._show_info("Message Sent", f"{channel_name} message sent successfully.")
            except Exception as exc:
                self._set_status(f"Messaging: {channel_name} failed")
                self._show_error("Messaging Error", str(exc))

        threading.Thread(target=_task, daemon=True).start()

    def send_sms(self) -> None:
        sid = self.twilio_sid_var.get().strip()
        token = self.twilio_token_var.get().strip()
        from_number = self.twilio_sms_from_var.get().strip()
        to_number = self.sms_to_var.get().strip()

        if not all([sid, token, from_number, to_number]):
            messagebox.showerror("Validation Error", "Fill Twilio SID, Token, SMS From and SMS To.")
            return

        def _send(message_text: str) -> None:
            client = Client(sid, token)
            client.messages.create(body=message_text, from_=from_number, to=to_number)

        self._run_in_background(_send, "SMS")

    def send_whatsapp(self) -> None:
        sid = self.twilio_sid_var.get().strip()
        token = self.twilio_token_var.get().strip()
        from_number = self.twilio_whatsapp_from_var.get().strip()
        to_number = self.whatsapp_to_var.get().strip()

        if not all([sid, token, from_number, to_number]):
            messagebox.showerror(
                "Validation Error", "Fill Twilio SID, Token, WhatsApp From and WhatsApp To."
            )
            return

        def _wa(number: str) -> str:
            return number if number.startswith("whatsapp:") else f"whatsapp:{number}"

        def _send(message_text: str) -> None:
            client = Client(sid, token)
            client.messages.create(body=message_text, from_=_wa(from_number), to=_wa(to_number))

        self._run_in_background(_send, "WhatsApp")

    def send_telegram(self) -> None:
        raw_token = self.telegram_bot_token_var.get().strip()
        raw_chat_id = self.telegram_chat_id_var.get().strip()

        # Allow users to paste env-style values such as BOT_TOKEN=... and CHAT_ID=...
        bot_token = re.sub(r"^\s*(?:TELEGRAM_)?BOT_TOKEN\s*=\s*", "", unquote(raw_token), flags=re.IGNORECASE)
        bot_token = bot_token.strip().strip('"').strip("'").replace(" ", "")

        chat_id = re.sub(r"^\s*(?:TELEGRAM_)?CHAT_ID\s*=\s*", "", unquote(raw_chat_id), flags=re.IGNORECASE)
        chat_id = chat_id.strip().strip('"').strip("'").replace(" ", "")

        if not bot_token or not chat_id:
            messagebox.showerror("Validation Error", "Fill Telegram Bot Token and Telegram Chat ID.")
            return

        if not re.fullmatch(r"\d{6,}:[A-Za-z0-9_-]{20,}", bot_token):
            messagebox.showerror(
                "Validation Error",
                "Telegram Bot Token looks invalid. Paste only the value after '=' (example: 123456789:ABC...).",
            )
            return

        if not (chat_id.startswith("@") or re.fullmatch(r"-?\d+", chat_id)):
            messagebox.showerror(
                "Validation Error",
                "Telegram Chat ID must be numeric (example: 5087327068 or -100...) or start with '@'.",
            )
            return

        def _send(message_text: str) -> None:
            response = requests.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": chat_id, "text": message_text},
                timeout=20,
            )
            payload = response.json()
            if response.status_code >= 400:
                raise RuntimeError(payload.get("description", f"Telegram API error: HTTP {response.status_code}"))
            if not payload.get("ok"):
                raise RuntimeError(payload.get("description", "Telegram API returned an unknown error."))

        self._run_in_background(_send, "Telegram")

    def save_to_excel(self) -> None:
        vehicle_no = self.vehicle_no_var.get().strip()
        if not vehicle_no:
            messagebox.showerror("Validation Error", "Vehicle No cannot be empty.")
            return

        if self.gross_weight is None or self.tare_weight is None:
            messagebox.showerror("Validation Error", "Capture both Gross and Tare weights before saving.")
            return

        self.calculate_net_weight()

        self._ensure_workbook()
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        row = [
            int(self.serial_no_var.get()),
            vehicle_no,
            self.date_var.get(),
            self.time_var.get(),
            self.challan_var.get().strip(),
            self.customer_code_var.get().strip(),
            self.customer_name_var.get().strip(),
            self.product_code_var.get().strip(),
            self.product_name_var.get().strip(),
            self.source_code_var.get().strip(),
            self.source_name_var.get().strip(),
            self.desti_code_var.get().strip(),
            self.desti_name_var.get().strip(),
            self.transporter_code_var.get().strip(),
            self.transporter_name_var.get().strip(),
            self.gross_weight,
            self.tare_weight,
            self.net_weight,
        ]
        sheet.append(row)

        workbook.save(EXCEL_FILE)
        workbook.close()

        messagebox.showinfo("Saved", "Weighment record saved successfully.")
        self.serial_no_var.set(str(int(self.serial_no_var.get()) + 1))

    def clear_fields(self, increment_serial: bool = False) -> None:
        if increment_serial:
            self.serial_no_var.set(str(int(self.serial_no_var.get()) + 1))

        self.vehicle_no_var.set("")
        self.challan_var.set("")
        self.customer_code_var.set("")
        self.customer_name_var.set("")
        self.product_code_var.set("")
        self.product_name_var.set("")
        self.source_code_var.set("")
        self.source_name_var.set("")
        self.desti_code_var.set("")
        self.desti_name_var.set("")
        self.transporter_code_var.set("")
        self.transporter_name_var.set("")
        self.gross_weight = None
        self.tare_weight = None
        self.net_weight = None

        self.gross_var.set("-")
        self.tare_var.set("-")
        self.net_var.set("-")


def main() -> None:
    root = tk.Tk()
    app = WeighmentApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
